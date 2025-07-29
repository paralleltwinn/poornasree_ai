import os
import re
import aiofiles
import PyPDF2
import docx
import docx2txt
from typing import Dict, Optional, List
import logging
from datetime import datetime
import uuid

# Enhanced PDF processing libraries
try:
    import pdfplumber
    import pypdf
    ADVANCED_PDF_SUPPORT = True
    print("✅ Advanced PDF support libraries loaded successfully!")
except ImportError as e:
    ADVANCED_PDF_SUPPORT = False
    print(f"⚠️ Advanced PDF support libraries not available: {e}. Using basic PDF extraction.")

# Optional Excel support
try:
    import openpyxl
    import xlrd
    import pandas as pd
    EXCEL_SUPPORT = True
    print("✅ Excel support libraries loaded successfully!")
except ImportError as e:
    EXCEL_SUPPORT = False
    print(f"⚠️ Excel support libraries not available: {e}. .xlsx/.xls files will not be supported.")

logger = logging.getLogger(__name__)

class DocumentService:
    """Service for processing and managing documents"""
    
    def __init__(self):
        self.upload_dir = "./data/uploads"
        # Set supported formats based on available libraries
        self.supported_formats = [".pdf", ".docx", ".doc", ".txt"]
        if EXCEL_SUPPORT:
            self.supported_formats.extend([".xlsx", ".xls"])
        self._ensure_upload_directory()
    
    def _ensure_upload_directory(self):
        """Create upload directory if it doesn't exist"""
        os.makedirs(self.upload_dir, exist_ok=True)
    
    async def save_uploaded_file(self, file_content: bytes, filename: str) -> str:
        """Save uploaded file to disk"""
        try:
            # Generate unique filename to avoid conflicts
            file_extension = os.path.splitext(filename)[1]
            unique_filename = f"{uuid.uuid4()}{file_extension}"
            file_path = os.path.join(self.upload_dir, unique_filename)
            
            async with aiofiles.open(file_path, 'wb') as f:
                await f.write(file_content)
            
            logger.info(f"File saved: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error saving file: {e}")
            raise e
    
    async def extract_text_from_file(self, file_path: str) -> Dict:
        """Extract text from various file formats"""
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            extracted_text = ""
            
            if file_extension == ".pdf":
                extracted_text = await self._extract_from_pdf(file_path)
            elif file_extension in [".docx", ".doc"]:
                extracted_text = await self._extract_from_docx(file_path)
            elif file_extension == ".txt":
                extracted_text = await self._extract_from_txt(file_path)
            elif file_extension in [".xlsx", ".xls"]:
                if EXCEL_SUPPORT:
                    extracted_text = await self._extract_from_excel(file_path)
                else:
                    raise ValueError(f"Excel support not available. Install openpyxl and pandas.")
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            return {
                "text": extracted_text,
                "word_count": len(extracted_text.split()),
                "char_count": len(extracted_text),
                "status": "success"
            }
            
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            return {
                "text": "",
                "word_count": 0,
                "char_count": 0,
                "status": "error",
                "error": str(e)
            }
    
    async def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file with enhanced extraction methods"""
        try:
            text = ""
            
            # Try advanced PDF processing first for better text extraction
            if ADVANCED_PDF_SUPPORT:
                try:
                    with pdfplumber.open(file_path) as pdf:
                        for page_num, page in enumerate(pdf.pages):
                            # Extract text with layout preservation
                            page_text = page.extract_text(layout=True)
                            if page_text:
                                text += f"\n--- PAGE {page_num + 1} ---\n"
                                text += page_text + "\n"
                            
                            # Also extract tables if they exist
                            tables = page.extract_tables()
                            if tables:
                                text += "\n--- TABLES ---\n"
                                for table_num, table in enumerate(tables):
                                    text += f"Table {table_num + 1}:\n"
                                    for row in table:
                                        if row and any(cell for cell in row if cell):
                                            clean_row = [str(cell).strip() if cell else "" for cell in row]
                                            text += " | ".join(clean_row) + "\n"
                                    text += "\n"
                    
                    if text.strip():
                        logger.info(f"Successfully extracted PDF text using pdfplumber: {len(text)} characters")
                        return self._enhance_pdf_text(text.strip())
                        
                except Exception as e:
                    logger.warning(f"pdfplumber extraction failed, trying pypdf: {e}")
                    
                # Fallback to pypdf for better text extraction than PyPDF2
                try:
                    with open(file_path, 'rb') as file:
                        pdf_reader = pypdf.PdfReader(file)
                        
                        for page_num, page in enumerate(pdf_reader.pages):
                            page_text = page.extract_text()
                            if page_text:
                                text += f"\n--- PAGE {page_num + 1} ---\n"
                                text += page_text + "\n"
                    
                    if text.strip():
                        logger.info(f"Successfully extracted PDF text using pypdf: {len(text)} characters")
                        return self._enhance_pdf_text(text.strip())
                        
                except Exception as e:
                    logger.warning(f"pypdf extraction failed, falling back to PyPDF2: {e}")
            
            # Fallback to original PyPDF2 method
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text:
                        text += f"\n--- PAGE {page_num + 1} ---\n"
                        text += page_text + "\n"
            
            if text.strip():
                logger.info(f"Successfully extracted PDF text using PyPDF2: {len(text)} characters")
                return self._enhance_pdf_text(text.strip())
            else:
                logger.warning("No text extracted from PDF file")
                return ""
            
        except Exception as e:
            logger.error(f"Error reading PDF: {e}")
            raise e
    
    async def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from Word document"""
        try:
            # Try using docx2txt first (more reliable for some files)
            text = docx2txt.process(file_path)
            
            if not text.strip():
                # Fallback to python-docx
                doc = docx.Document(file_path)
                text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
            return text.strip()
            
        except Exception as e:
            logger.error(f"Error reading DOCX: {e}")
            raise e
    
    async def _extract_from_txt(self, file_path: str) -> str:
        """Extract text from text file"""
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8') as file:
                text = await file.read()
            return text.strip()
            
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                async with aiofiles.open(file_path, 'r', encoding='latin-1') as file:
                    text = await file.read()
                return text.strip()
            except Exception as e:
                logger.error(f"Error reading text file with fallback encoding: {e}")
                raise e
        except Exception as e:
            logger.error(f"Error reading text file: {e}")
            raise e
    
    async def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file with multiple sheets support"""
        if not EXCEL_SUPPORT:
            raise ValueError("Excel support not available. Install openpyxl and pandas.")
            
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            extracted_content = []
            
            if file_extension == ".xlsx":
                # Use openpyxl for .xlsx files
                try:
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        sheet_content = [f"\n=== SHEET: {sheet_name} ===\n"]
                        
                        # Extract all cell values from the sheet
                        for row in sheet.iter_rows(values_only=True):
                            row_data = []
                            for cell_value in row:
                                if cell_value is not None:
                                    # Convert to string and clean up
                                    cell_str = str(cell_value).strip()
                                    if cell_str:
                                        row_data.append(cell_str)
                            
                            if row_data:  # Only add non-empty rows
                                sheet_content.append(" | ".join(row_data))
                        
                        if len(sheet_content) > 1:  # Has content beyond header
                            extracted_content.extend(sheet_content)
                            extracted_content.append("\n")  # Add separator between sheets
                    
                    workbook.close()
                    
                except Exception as e:
                    logger.warning(f"Failed to read .xlsx with openpyxl, trying pandas: {e}")
                    # Fallback to pandas for .xlsx
                    extracted_content = await self._extract_excel_with_pandas(file_path)
                    
            elif file_extension == ".xls":
                # Use pandas for .xls files (better compatibility)
                extracted_content = await self._extract_excel_with_pandas(file_path)
            
            # Join all content
            final_text = "\n".join(extracted_content)
            
            # Clean up the text
            final_text = self._clean_excel_text(final_text)
            
            logger.info(f"Successfully extracted text from Excel file: {len(final_text)} characters")
            return final_text
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise e
    
    async def _extract_excel_with_pandas(self, file_path: str) -> List[str]:
        """Extract Excel content using pandas (handles both .xlsx and .xls)"""
        if not EXCEL_SUPPORT:
            raise ValueError("Excel support not available.")
            
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            extracted_content = []
            
            for sheet_name in excel_file.sheet_names:
                # Read each sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                
                # Convert to text
                sheet_content = [f"\n=== SHEET: {sheet_name} ===\n"]
                
                for index, row in df.iterrows():
                    row_data = []
                    for value in row:
                        if pd.notna(value) and str(value).strip():
                            row_data.append(str(value).strip())
                    
                    if row_data:  # Only add non-empty rows
                        sheet_content.append(" | ".join(row_data))
                
                if len(sheet_content) > 1:  # Has content beyond header
                    extracted_content.extend(sheet_content)
                    extracted_content.append("\n")  # Add separator between sheets
            
            return extracted_content
            
        except Exception as e:
            logger.error(f"Error reading Excel with pandas: {e}")
            raise e
    
    async def extract_excel_rows(self, file_path: str) -> List[Dict]:
        """Extract Excel data row by row for service guide training"""
        if not EXCEL_SUPPORT:
            raise ValueError("Excel support not available. Install openpyxl and pandas.")
        
        try:
            # Read Excel file with all sheets
            df_dict = pd.read_excel(file_path, sheet_name=None)
            all_rows = []
            
            for sheet_name, df in df_dict.items():
                # Clean the dataframe
                df = df.dropna(how='all')  # Remove completely empty rows
                df = df.fillna('')  # Fill NaN with empty strings
                
                # Get headers
                headers = [str(col).strip() for col in df.columns]
                
                # Extract each row
                for index, row in df.iterrows():
                    row_data = {
                        "sheet": sheet_name,
                        "row_number": index + 2,  # +2 for Excel numbering (1-indexed + header)
                        "columns": {}
                    }
                    
                    # Map each column to its value
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            value = str(row.iloc[col_idx]).strip()
                            if value and value != 'nan':
                                row_data["columns"][header] = value
                    
                    # Only include rows with data
                    if row_data["columns"]:
                        all_rows.append(row_data)
            
            logger.info(f"Extracted {len(all_rows)} rows from Excel file")
            return all_rows
            
        except Exception as e:
            logger.error(f"Error extracting Excel rows: {e}")
            raise e

    def _clean_excel_text(self, text: str) -> str:
        """Clean up extracted Excel text"""
        # Remove excessive whitespace
        text = re.sub(r'\n\s*\n\s*\n', '\n\n', text)
        
        # Remove empty separator lines
        text = re.sub(r'\n\s*\|\s*\n', '\n', text)
        
        # Clean up pipe separators
        text = re.sub(r'\s*\|\s*', ' | ', text)
        
        # Remove leading/trailing whitespace
        text = text.strip()
        
        return text
    
    def validate_file(self, filename: str, file_size: int) -> Dict:
        """Validate uploaded file"""
        # Check file extension
        file_extension = os.path.splitext(filename)[1].lower()
        if file_extension not in self.supported_formats:
            return {
                "valid": False,
                "error": f"Unsupported file format. Supported formats: {', '.join(self.supported_formats)}"
            }
        
        # Check file size (max 10MB for free tier)
        max_size = 10 * 1024 * 1024  # 10MB
        if file_size > max_size:
            return {
                "valid": False,
                "error": f"File too large. Maximum size: {max_size // (1024*1024)}MB"
            }
        
        return {"valid": True}
    
    async def process_document(self, file_path: str, filename: str) -> Dict:
        """Process document and prepare for AI service with enhanced text processing"""
        try:
            # Extract text
            extraction_result = await self.extract_text_from_file(file_path)
            
            if extraction_result["status"] != "success":
                return extraction_result
            
            # Enhanced text preprocessing with document identification
            raw_text = extraction_result["text"]
            
            # Add document identification wrapper for better AI recognition
            file_extension = os.path.splitext(filename)[1].lower()
            document_identifier = f"=== DOCUMENT: {filename} ===\n"
            document_identifier += f"=== FORMAT: {file_extension.upper().replace('.', '')} ===\n"
            document_identifier += f"=== UPLOAD_TIME: {datetime.now().isoformat()} ===\n\n"
            
            processed_text = self._enhance_text_processing(raw_text)
            
            # Wrap processed text with clear document boundaries
            final_text = document_identifier + processed_text + f"\n\n=== END OF {filename} ==="
            
            # Prepare enhanced metadata
            metadata = {
                "filename": filename,
                "source": filename,
                "upload_date": datetime.now().isoformat(),
                "file_path": file_path,
                "raw_word_count": extraction_result["word_count"],
                "raw_char_count": extraction_result["char_count"],
                "processed_word_count": len(final_text.split()),
                "processed_char_count": len(final_text),
                "document_type": self._detect_document_type(processed_text),
                "key_sections": self._extract_key_sections(processed_text),
                "processing_version": "2.1",
                "enhancement_features": ["document_identification", "text_structuring", "advanced_pdf_extraction"]
            }
            
            return {
                "text": final_text,
                "metadata": metadata,
                "status": "success",
                "processing_stats": {
                    "original_length": len(raw_text),
                    "processed_length": len(final_text),
                    "improvement_ratio": len(final_text) / len(raw_text) if len(raw_text) > 0 else 0
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing document: {e}")
            return {
                "text": "",
                "metadata": {},
                "status": "error",
                "error": str(e)
            }
    
    def _enhance_text_processing(self, text: str) -> str:
        """Enhanced text processing for better AI training"""
        # Remove excessive whitespace but preserve structure
        text = re.sub(r'[ \t]+', ' ', text)  # Multiple spaces/tabs to single space
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)  # Multiple newlines to double newline
        
        # Fix common OCR errors
        text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)  # Add space between words
        text = re.sub(r'(\d+)([A-Za-z])', r'\1 \2', text)  # Add space between numbers and letters
        text = re.sub(r'([A-Za-z])(\d+)', r'\1 \2', text)  # Add space between letters and numbers
        
        # Standardize punctuation
        text = re.sub(r'["""]', '"', text)  # Normalize quotes
        text = re.sub(r"[''']", "'", text)  # Normalize apostrophes
        text = re.sub(r'[—–]', '-', text)  # Normalize dashes
        
        # Preserve section headers and lists
        lines = text.split('\n')
        processed_lines = []
        
        for line in lines:
            line = line.strip()
            if line:
                # Detect and preserve section headers
                if self._is_section_header(line):
                    processed_lines.append(f"\n## {line}\n")
                # Detect and preserve list items
                elif self._is_list_item(line):
                    processed_lines.append(f"• {line}")
                # Detect and preserve procedure steps
                elif self._is_procedure_step(line):
                    processed_lines.append(f"🔧 {line}")
                # Detect and preserve safety warnings
                elif self._is_safety_warning(line):
                    processed_lines.append(f"⚠️ {line}")
                else:
                    processed_lines.append(line)
        
        return '\n'.join(processed_lines).strip()
    
    def _is_section_header(self, line: str) -> bool:
        """Detect if a line is a section header"""
        line_lower = line.lower()
        return (
            len(line.split()) <= 8 and  # Short lines
            (line.isupper() or  # All caps
             any(word in line_lower for word in ['chapter', 'section', 'part', 'introduction', 'overview', 'procedure', 'maintenance', 'safety', 'operation']) or
             re.match(r'^\d+\.?\s+[A-Z]', line))  # Numbered sections
        )
    
    def _is_list_item(self, line: str) -> bool:
        """Detect if a line is a list item"""
        return (
            re.match(r'^\s*[•\-\*]\s+', line) or  # Bullet points
            re.match(r'^\s*[a-zA-Z]\.\s+', line) or  # Letter lists
            re.match(r'^\s*\d+\.\s+', line)  # Numbered lists
        )
    
    def _is_procedure_step(self, line: str) -> bool:
        """Detect if a line is a procedure step"""
        line_lower = line.lower()
        return any(word in line_lower for word in [
            'step', 'first', 'then', 'next', 'finally', 'before', 'after',
            'turn on', 'turn off', 'press', 'push', 'pull', 'rotate', 'adjust'
        ])
    
    def _is_safety_warning(self, line: str) -> bool:
        """Detect if a line is a safety warning"""
        line_lower = line.lower()
        return any(word in line_lower for word in [
            'warning', 'caution', 'danger', 'notice', 'important', 'critical',
            'never', 'always', 'must', 'required', 'mandatory'
        ])
    
    def _detect_document_type(self, text: str) -> str:
        """Detect the type of document based on content"""
        text_lower = text.lower()
        
        # Check for Excel-specific content
        if '=== sheet:' in text_lower:
            if any(word in text_lower for word in ['inventory', 'stock', 'parts', 'components']):
                return 'inventory_spreadsheet'
            elif any(word in text_lower for word in ['schedule', 'timeline', 'calendar', 'date']):
                return 'schedule_spreadsheet'
            elif any(word in text_lower for word in ['maintenance', 'service', 'checklist']):
                return 'maintenance_spreadsheet'
            elif any(word in text_lower for word in ['cost', 'price', 'budget', 'expense']):
                return 'financial_spreadsheet'
            elif any(word in text_lower for word in ['specification', 'specs', 'parameters']):
                return 'specification_spreadsheet'
            else:
                return 'data_spreadsheet'
        
        # Existing document type detection
        if any(word in text_lower for word in ['manual', 'instruction', 'guide', 'handbook']):
            if any(word in text_lower for word in ['maintenance', 'service', 'repair']):
                return 'maintenance_manual'
            elif any(word in text_lower for word in ['operation', 'operating', 'user']):
                return 'operation_manual'
            elif any(word in text_lower for word in ['installation', 'setup', 'assembly']):
                return 'installation_guide'
            else:
                return 'general_manual'
        elif any(word in text_lower for word in ['safety', 'warning', 'hazard']):
            return 'safety_document'
        elif any(word in text_lower for word in ['specification', 'spec', 'technical']):
            return 'technical_specification'
        elif any(word in text_lower for word in ['troubleshoot', 'problem', 'error', 'fault']):
            return 'troubleshooting_guide'
        else:
            return 'general_document'
    
    def _extract_key_sections(self, text: str) -> List[str]:
        """Extract key sections from the document"""
        sections = []
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            # Handle Excel sheet headers
            if line.startswith('=== SHEET:') and len(line) > 10:
                sheet_name = line.replace('=== SHEET:', '').replace('===', '').strip()
                sections.append(f"Sheet: {sheet_name}")
            # Handle regular section headers
            elif line.startswith('## ') and len(line) > 5:
                section_name = line[3:].strip()
                sections.append(section_name)
        
        return sections[:15]  # Increased limit for Excel files with multiple sheets
    
    def cleanup_file(self, file_path: str) -> bool:
        """Clean up temporary files"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Cleaned up file: {file_path}")
                return True
            return False
        except Exception as e:
            logger.error(f"Error cleaning up file {file_path}: {e}")
            return False
    
    def get_supported_formats(self) -> list:
        """Get list of supported file formats"""
        return self.supported_formats
    
    def get_upload_stats(self) -> Dict:
        """Get upload directory statistics"""
        try:
            files = os.listdir(self.upload_dir)
            total_files = len(files)
            total_size = sum(os.path.getsize(os.path.join(self.upload_dir, f)) for f in files)
            
            return {
                "total_files": total_files,
                "total_size_mb": round(total_size / (1024 * 1024), 2),
                "upload_directory": self.upload_dir
            }
        except Exception as e:
            logger.error(f"Error getting upload stats: {e}")
            return {
                "total_files": 0,
                "total_size_mb": 0,
                "upload_directory": self.upload_dir,
                "error": str(e)
            }
    
    def _enhance_pdf_text(self, text: str) -> str:
        """Enhance PDF text extraction with better formatting and structure recognition"""
        try:
            # Split into lines for processing
            lines = text.split('\n')
            enhanced_lines = []
            current_chapter = ""
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Detect and format headers/chapters
                if re.match(r'^(chapter|section|part)\s+\d+', line, re.IGNORECASE):
                    current_chapter = line.upper()
                    enhanced_lines.append(f"\n=== {current_chapter} ===")
                    continue
                
                # Detect chapter headers in content
                if any(word in line.lower() for word in ['chapter 1:', 'chapter 2:', 'chapter 3:', 'chapter 4:', 'chapter 5:', 'chapter 6:']):
                    current_chapter = line.upper()
                    enhanced_lines.append(f"\n=== {current_chapter} ===")
                    continue
                
                # Detect numbered lists and procedures
                if re.match(r'^\d+[\.\)]\s+', line):
                    enhanced_lines.append(f"PROCEDURE_STEP: {line}")
                    continue
                
                # Detect bullet points
                if re.match(r'^[•\-\*]\s+', line):
                    enhanced_lines.append(f"REQUIREMENT: {line[1:].strip()}")
                    continue
                
                # Detect specifications (key: value pairs)
                if ':' in line and len(line.split(':')) == 2:
                    key, value = line.split(':', 1)
                    if len(key.strip()) < 50 and len(value.strip()) < 100:
                        enhanced_lines.append(f"SPECIFICATION: {key.strip()}: {value.strip()}")
                        continue
                
                # Detect error codes with more context
                if re.search(r'\b[E]\d{3,4}\b', line, re.IGNORECASE):
                    enhanced_lines.append(f"ERROR_CODE_INFO: {line}")
                    # Add context if this is about error meaning
                    if any(word in line.lower() for word in ['cause', 'solution', 'indicates', 'means']):
                        enhanced_lines.append(f"ERROR_EXPLANATION: {line}")
                    continue
                
                # Detect model numbers and technical identifiers with more prominence
                if re.search(r'\b(PMC|CNC|MODEL)\W*\d+', line, re.IGNORECASE):
                    enhanced_lines.append(f"MACHINE_MODEL: {line}")
                    continue
                
                # Detect safety instructions with more prominence
                if any(word in line.lower() for word in ['safety', 'warning', 'caution', 'danger', 'protective', 'equipment']):
                    enhanced_lines.append(f"SAFETY_INSTRUCTION: {line}")
                    continue
                
                # Detect maintenance instructions with more detail
                if any(word in line.lower() for word in ['maintenance', 'lubricate', 'inspect', 'replace', 'calibrate']):
                    enhanced_lines.append(f"MAINTENANCE_TASK: {line}")
                    # Add frequency context
                    if any(word in line.lower() for word in ['hours', 'daily', 'weekly', 'monthly', 'quarterly']):
                        enhanced_lines.append(f"MAINTENANCE_SCHEDULE: {line}")
                    continue
                
                # Detect technical specifications in tables
                if any(indicator in line.lower() for indicator in ['spindle', 'axis', 'travel', 'accuracy', 'capacity', 'power', 'speed']):
                    enhanced_lines.append(f"TECHNICAL_SPEC: {line}")
                    continue
                
                # Detect warranty and support information
                if any(word in line.lower() for word in ['warranty', 'years', 'support', 'training', 'course']):
                    enhanced_lines.append(f"SUPPORT_INFO: {line}")
                    continue
                
                # Regular line with chapter context
                if current_chapter:
                    enhanced_lines.append(f"[{current_chapter}] {line}")
                else:
                    enhanced_lines.append(line)
            
            # Join with preserved structure
            enhanced_text = '\n'.join(enhanced_lines)
            
            # Add comprehensive document metadata for better AI recognition
            doc_header = "=== PMC-2000 CNC MACHINE MANUAL ===\n"
            doc_header += "=== ENHANCED PDF DOCUMENT ===\n"
            doc_header += "=== CONTENT: TECHNICAL SPECIFICATIONS, SAFETY, PROCEDURES, MAINTENANCE ===\n\n"
            
            enhanced_text = doc_header + enhanced_text + "\n\n=== END PMC-2000 MANUAL ==="
            
            return enhanced_text
            
        except Exception as e:
            logger.warning(f"Error enhancing PDF text: {e}")
            return text
