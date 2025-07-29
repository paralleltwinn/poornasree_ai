#!/usr/bin/env python3
"""
Create a test Excel file with multiple sheets for testing
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import os

def create_test_excel():
    """Create a test Excel file with multiple sheets containing CNC machine data"""
    print("🔧 Creating test Excel file with multiple sheets...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Sheet 1: Machine Specifications
    specs_sheet = wb.create_sheet("Machine_Specifications")
    specs_data = [
        ["Parameter", "Value", "Unit", "Notes"],
        ["Max Spindle Speed", "10000", "RPM", "Variable frequency drive"],
        ["Table Size", "1200x800", "mm", "T-slot aluminum"],
        ["Travel X-Axis", "1000", "mm", "Ball screw drive"],
        ["Travel Y-Axis", "600", "mm", "Ball screw drive"],
        ["Travel Z-Axis", "500", "mm", "Ball screw drive"],
        ["Tool Capacity", "20", "tools", "Automatic tool changer"],
        ["Coolant Capacity", "100", "liters", "Flood coolant system"],
        ["Power Requirement", "15", "kW", "3-phase 380V"],
        ["Positioning Accuracy", "0.01", "mm", "Repeatability ±0.005mm"],
        ["Control System", "Fanuc 0i-MF", "", "Latest generation controller"]
    ]
    
    for row_num, row_data in enumerate(specs_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = specs_sheet.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 2: Maintenance Schedule
    maintenance_sheet = wb.create_sheet("Maintenance_Schedule")
    maintenance_data = [
        ["Task", "Frequency", "Duration", "Responsible", "Last Done", "Next Due"],
        ["Lubricate ball screws", "Weekly", "30 min", "Operator", "2025-07-21", "2025-07-28"],
        ["Check coolant level", "Daily", "5 min", "Operator", "2025-07-27", "2025-07-28"],
        ["Clean chip tray", "Daily", "15 min", "Operator", "2025-07-27", "2025-07-28"],
        ["Inspect tool holder", "Monthly", "2 hours", "Technician", "2025-07-01", "2025-08-01"],
        ["Calibrate axes", "Quarterly", "4 hours", "Engineer", "2025-04-01", "2025-10-01"],
        ["Replace air filters", "Monthly", "1 hour", "Technician", "2025-07-01", "2025-08-01"],
        ["Check spindle bearings", "Annually", "8 hours", "Specialist", "2025-01-15", "2026-01-15"],
        ["Update software", "As needed", "2 hours", "Engineer", "2025-06-15", "TBD"],
        ["Full machine inspection", "Annually", "16 hours", "Specialist", "2025-01-15", "2026-01-15"]
    ]
    
    for row_num, row_data in enumerate(maintenance_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = maintenance_sheet.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 3: Tool Inventory
    tools_sheet = wb.create_sheet("Tool_Inventory")
    tools_data = [
        ["Tool Number", "Type", "Diameter", "Material", "Location", "Condition", "Notes"],
        ["T01", "End Mill", "10mm", "Carbide", "Magazine-1", "Good", "General purpose"],
        ["T02", "End Mill", "6mm", "Carbide", "Magazine-2", "Good", "Fine finishing"],
        ["T03", "Drill", "8mm", "HSS", "Magazine-3", "Fair", "Replace soon"],
        ["T04", "Face Mill", "50mm", "Carbide", "Magazine-4", "Excellent", "Heavy cutting"],
        ["T05", "Tap", "M8", "HSS", "Magazine-5", "Good", "Threading"],
        ["T06", "Reamer", "12mm", "HSS", "Magazine-6", "Good", "Precision holes"],
        ["T07", "Chamfer", "45°", "Carbide", "Magazine-7", "Good", "Edge preparation"],
        ["T08", "Ball End", "8mm", "Carbide", "Magazine-8", "Excellent", "3D machining"],
        ["T09", "Slot Drill", "4mm", "Carbide", "Magazine-9", "Good", "Slot cutting"],
        ["T10", "Boring Bar", "20mm", "Carbide", "Magazine-10", "Good", "Internal boring"]
    ]
    
    for row_num, row_data in enumerate(tools_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = tools_sheet.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Sheet 4: Error Codes
    errors_sheet = wb.create_sheet("Error_Codes")
    errors_data = [
        ["Code", "Description", "Cause", "Solution"],
        ["E001", "Spindle Overheat", "Excessive cutting speed", "Reduce spindle speed, check coolant"],
        ["E002", "Axis Following Error", "Mechanical binding", "Check for obstacles, lubricate"],
        ["E003", "Tool Break Detection", "Tool broken during cut", "Replace tool, check cutting parameters"],
        ["E004", "Emergency Stop", "Safety circuit activated", "Clear safety condition, reset"],
        ["E005", "Coolant Low Level", "Insufficient coolant", "Refill coolant tank"],
        ["E006", "Air Pressure Low", "Compressor issue", "Check air compressor, filters"],
        ["E007", "Power Failure", "Electrical supply issue", "Check main power supply"],
        ["E008", "Memory Full", "Program storage limit", "Delete old programs, backup data"],
        ["E009", "Tool Magazine Error", "ATC malfunction", "Check tool changer mechanism"],
        ["E010", "Servo Fault", "Motor drive error", "Check servo connections, reset drive"]
    ]
    
    for row_num, row_data in enumerate(errors_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = errors_sheet.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save the file
    file_path = "test_cnc_data.xlsx"
    wb.save(file_path)
    
    print(f"✅ Test Excel file created: {file_path}")
    print(f"   Sheets: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"   - {sheet_name}: {sheet.max_row} rows x {sheet.max_column} columns")
    
    return file_path

if __name__ == "__main__":
    create_test_excel()
